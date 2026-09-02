from pathlib import Path
from datetime import datetime, date
import logging, math
import pandas as pd
from openpyxl import load_workbook
import re

log = logging.getLogger(__name__)

REQUIRED_SHEETS = [
    "Cuadro de Mando", "Comp 2025-2026", "Comp PP 2025-2026",
    "Comp Aportes", "80-20", "Oportunidad Crecimiento", "Oport Crecimiento Zona",
]
OPTIONAL_SHEETS = ["Comp 2024-2025", "Oport Crecimiento Zona", "2024", "2025", "2026",
                   "Periodo Parcial 2025", "Periodo Parcial 2026", "Hoja1"]

def clean_value(v):
    if v is None: return None
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)): return None
    if isinstance(v, (datetime, date)): return v.isoformat()
    try:
        if hasattr(v, "item"): return v.item()
    except Exception: pass
    return v

def worksheet_to_df(ws):
    values = list(ws.values)
    if not values: return pd.DataFrame()
    # Find the first row with useful text/numeric headers. Preserve sheets with title rows.
    header_idx = 0
    for i, row in enumerate(values[:12]):
        nonempty = [x for x in row if x is not None]
        if len(nonempty) >= 2:
            header_idx = i
            # Prefer a row containing known business fields
            txt = " ".join(str(x).lower() for x in nonempty)
            if any(k in txt for k in ["rut", "empresa", "zona", "holding", "aporte", "ejecutiva", "dif aporte", "imponible"]):
                break
    headers = list(values[header_idx])
    # make unique headers
    seen = {}
    clean_headers = []
    for j,h in enumerate(headers):
        name = str(h).strip() if h is not None else f"Columna_{j+1}"
        if not name or name.lower() == "none": name = f"Columna_{j+1}"
        seen[name] = seen.get(name,0)+1
        clean_headers.append(name if seen[name]==1 else f"{name}_{seen[name]}")
    rows = values[header_idx+1:]
    df = pd.DataFrame(rows, columns=clean_headers)
    df = df.dropna(how="all")
    return df

class ExcelProcessor:
    def __init__(self, path: Path):
        self.path = Path(path)
        self.last_reload = None
        self.warnings = []
        self.sheets = []
        self.frames = {}
        self.meta = {}

    def load(self):
        self.warnings = []
        if not self.path.exists():
            self.warnings = [f"No se encontró el archivo: {self.path}"]
            log.warning(self.warnings[0]); return self

        # Formula view validates the XLSM structure while data_only reads Excel's cached results.
        wb_formula = load_workbook(self.path, keep_vba=True, data_only=False, read_only=True)
        self.sheets = wb_formula.sheetnames
        wb_formula.close()
        wb_values = load_workbook(self.path, keep_vba=True, data_only=True, read_only=True)
        for sheet in self.sheets:
            try:
                self.frames[sheet] = worksheet_to_df(wb_values[sheet])
                log.info("Hoja %s: %s registros", sheet, len(self.frames[sheet]))
            except Exception as exc:
                self.warnings.append(f"No fue posible leer '{sheet}': {exc}")
                log.exception("Error leyendo %s", sheet)
        wb_values.close()

        missing = [s for s in REQUIRED_SHEETS if s not in self.sheets]
        self.warnings.extend([f"Falta la hoja '{s}'" for s in missing])
        self.meta["mes_comparacion"] = self._cell("Comp PP 2025-2026", 1, 2)
        self.last_reload = datetime.now().isoformat()
        return self

    def _cell(self, sheet, row, col):
        if not self.path.exists(): return None
        wb = load_workbook(self.path, data_only=True, read_only=True, keep_vba=True)
        value = wb[sheet].cell(row, col).value if sheet in wb.sheetnames else None
        wb.close()
        return clean_value(value)

    def _records(self, sheet):
        df = self.frames.get(sheet, pd.DataFrame()).copy()
        return [{str(k): clean_value(v) for k,v in row.items()} for row in df.to_dict(orient="records")]

    def summary(self):
        return {"last_update": self.last_reload, "sheets": self.sheets,
                "warnings": self.warnings, "records": {k: len(v) for k,v in self.frames.items()},
                "mes_comparacion": self.meta.get("mes_comparacion")}

    def process_cuadro_mando(self): return {"rows": self._records("Cuadro de Mando")}
    def process_comp_2025_2026(self): return {"rows": self._records("Comp 2025-2026")}
    def process_comp_pp(self): return {"rows": self._records("Comp PP 2025-2026"), "mes_comparacion": self.meta.get("mes_comparacion")}
    def process_comp_aportes(self): return {"rows": self._records("Comp Aportes")}
    def process_80_20(self): return {"rows": self._records("80-20")}
    def process_oportunidad(self): return {"rows": self._records("Oportunidad Crecimiento")}
    def process_oportunidad_zona(self): return {"rows": self._records("Oport Crecimiento Zona")}
    def reload(self):
        self.frames = {}; return self.load()
